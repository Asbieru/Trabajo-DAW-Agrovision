"""
main.py  -  Servidor Flask  (AgroVision · bd_proyectofinal)
"""


from flask import Flask, render_template, request, redirect, url_for, abort, jsonify, make_response, session
from functools import wraps
import jwt
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from usuarioAD import (autenticarUsuario, buscarUsuarioPorCorreo, obtenerUsuarios,
                       listarUsuariosCompleto, obtenerPerfilUsuario,
                       estadisticasUsuario, historialParticipacionUsuario,
                       resumenHistorialUsuario, insertarUsuario)
from ticketAD import (Ticket, listarTickets, insertarTicket, obtenerTicket,
                      resolverTicket, guardarCalificacionTicket,
                      editarTicket, cancelarTicket, listarAplicaciones,
                      listarAplicacionesActivas,
                      listarPosiblesAgentes, asignarTicket,
                      reasignarTicket, reabrirTicket,
                      listarUsuariosNivelMenor,
                      obtenerAplicacion, calcularSLA,
                      obtenerDetallesTicket,
                      obtenerCalificacionTicket,
                      insertarAplicacion, editarAplicacion, eliminarAplicacion,
                      toggleEstadoAplicacion)
from actividadAD import (Actividad, listarActividades, insertarActividad,
                         actualizarEstadoActividad, eliminarActividad,
                         desbloquearActividad,
                         listarTodosSprints, listarAsignadosPorProyecto,
                         resumenActividadesPorProyecto, proximoCodigo)
from proyectoAD import (Proyecto, listarProyectos, insertarProyecto,
                        obtenerProyecto, actualizarProyecto,
                        listarAvances, insertarAvance, eliminarAvance,
                        eliminarProyecto, tieneActividadesPendientes,
                        listarProyectosEnRevision, aprobarProyecto, rechazarProyecto,
                        generarSprintsProyecto, listarSprintsPorProyecto)
from indicadoresAD import (resumenKPI, kpiPorAplicacion, kpiPorPrioridad,
                            kpiPorAgente, kpiPorMes, kpiSprintsActivos,
                            kpiSatisfaccion, comentariosCalificacionesRecientes,
                            kpiProyectosPorEstado, kpiVelocityPorSprint,
                            kpiCargaPorProgramador, kpiTiempoRespuesta,
                            kpiPorIntensidad, kpiSLAPorAgente,
                            kpiProyectosPorSalud, kpiAvancePromedioPorProyecto,
                            kpiTiempoResolucionPorAplicacion,
                            kpiCancelados, kpiRankingAppsProblemáticas,
                            kpiPorTipo, kpiTop5MasLentos,
                            kpiActividadesPorEstado, kpiProgramadoresSinCarga,
                            kpiProyectosVencidos)
from reportesAD import (reporteResumen, reporteTicketsPorApp, reporteTicketsPorTipo,
                         reporteStoryPointsPorProgramador, reporteCarryoverPorProgramador,
                         reporteTicketsFiltrados, obtenerAplicaciones,
                         reporteProyectosPorEstado, reporteProyectosEnRiesgo,
                         reporteRendimientoPorSprint, reporteProyectosFiltrados,
                         obtenerResponsables,
                         reporteActividadesPorProyecto, resumenActividadesReporte,
                         reporteSLAPorAplicacion, reporteTicketsPorEstado,
                         reporteAgentesMetricas, reporteTendenciaPorMes)
from rolAD import (listarRoles, obtenerRol, insertarRol, actualizarRol, eliminarRol)
from permisoAD import (listarPermisos, obtenerPermiso, insertarPermiso,
                       actualizarPermiso, eliminarPermiso, obtenerPermisosPorRol)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'agrovision-clave-secreta-2024'
app.debug = True


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('usuario'):
            if request.path.startswith('/api/'):
                return jsonify({'ok': False, 'mensaje': 'No autenticado.'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


JWT_SECRET = 'agrovision-clave-secreta-2024'
JWT_ALGORITHM = 'HS256'


def jwt_required(f=None):
    """Verifica el token JWT del header Authorization: Bearer <token>.
    Usar como @jwt_required o @jwt_required()."""
    if f is not None:
        # Usado como @jwt_required (sin paréntesis)
        @wraps(f)
        def decorated(*args, **kwargs):
            auth = request.headers.get('Authorization', '')
            if auth.startswith('Bearer '):
                token = auth[7:]
                try:
                    payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
                    if not session.get('usuario'):
                        session['usuario'] = payload['usuario']
                        session.permanent = True
                    return f(*args, **kwargs)
                except jwt.ExpiredSignatureError:
                    return jsonify({'ok': False, 'mensaje': 'Token expirado.'}), 401
                except jwt.InvalidTokenError:
                    return jsonify({'ok': False, 'mensaje': 'Token inválido.'}), 401
            return f(*args, **kwargs)
        return decorated

    # Usado como @jwt_required() (con paréntesis)
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            auth = request.headers.get('Authorization', '')
            if auth.startswith('Bearer '):
                token = auth[7:]
                try:
                    payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
                    if not session.get('usuario'):
                        session['usuario'] = payload['usuario']
                        session.permanent = True
                    return f(*args, **kwargs)
                except jwt.ExpiredSignatureError:
                    return jsonify({'ok': False, 'mensaje': 'Token expirado.'}), 401
                except jwt.InvalidTokenError:
                    return jsonify({'ok': False, 'mensaje': 'Token inválido.'}), 401
            return f(*args, **kwargs)
        return decorated
    return decorator


@app.context_processor
def inject_usuario():
    return dict(usuario=session.get('usuario'))


# ──────────────────────────────────────────────────────────────
#  MANEJADORES DE ERROR
# ──────────────────────────────────────────────────────────────

@app.errorhandler(400)
def error_400(e):
    return render_template('error400.html'), 400

@app.errorhandler(404)
def error_404(e):
    return render_template('error404.html'), 404

@app.errorhandler(500)
def error_500(e):
    return render_template('error500.html'), 500

@app.route('/error-500')
def pagina_error_500():
    """Ruta directa para redirigir a la página de error 500 desde JS."""
    return render_template('error500.html'), 500

# ──────────────────────────────────────────────────────────────
#  RUTA RAIZ
# ──────────────────────────────────────────────────────────────

@app.route('/')
def raiz():
    return redirect(url_for('login'))


# ──────────────────────────────────────────────────────────────
#  LOGIN
# ──────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None

    if request.method == 'POST':
        correo   = request.form.get('correo', '').strip()
        password = request.form.get('password', '')

        try:
            ok, mensaje, datos = autenticarUsuario(correo, password)
        except Exception as e:
            print(f'Error de conexión al autenticar: {e}')
            abort(500)

        if ok:
            session['usuario'] = datos
            session.permanent = True
            return redirect(url_for('index'))
        else:
            error = mensaje

    return render_template('login.html', error=error)


# ──────────────────────────────────────────────────────────────
#  OLVIDÉ MI CONTRASEÑA
# ──────────────────────────────────────────────────────────────

@app.route('/olvide-contrasena', methods=['GET', 'POST'])
def olvide_contrasena():
    mensaje = None
    tipo    = None

    if request.method == 'POST':
        correo = request.form.get('correo', '').strip()
        try:
            encontrado = buscarUsuarioPorCorreo(correo)
        except Exception as e:
            print(f'Error de conexión al buscar correo: {e}')
            abort(500)

        if encontrado:
            mensaje = ('Si el correo está registrado, recibirás un enlace '
                       'para restablecer tu contraseña.')
            tipo = 'exito'
        else:
            mensaje = 'No se encontró una cuenta activa con ese correo.'
            tipo = 'error'

    return render_template('olvide_contrasena.html', mensaje=mensaje, tipo=tipo)


# ──────────────────────────────────────────────────────────────
#  LOGOUT
# ──────────────────────────────────────────────────────────────

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ──────────────────────────────────────────────────────────────
#  DASHBOARD
# ──────────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def index():
    return render_template('panelDeControl.html')


# ──────────────────────────────────────────────────────────────
#  TICKETS DE SOPORTE
# ──────────────────────────────────────────────────────────────

@app.route('/ticket/nuevo')
@login_required
def form_ticket():
    aplicaciones = listarAplicacionesActivas()
    return render_template('NuevoTicket.html', aplicaciones=aplicaciones)


@app.route('/tickets')
@login_required
def listar_tickets():
    try:
        tickets = listarTickets()
    except Exception as e:
        print(f'Error al listar tickets: {e}')
        abort(500)
    return render_template('gestionTicket.html', tickets=tickets)


@app.route('/tickets/resolver')
@login_required
def resolver_tickets():
    id_usuario = session['usuario']['id_usuario']
    todos = listarTickets()
    pendientes = [t for t in todos if t['estado'] == 'en_progreso'
                  and t.get('id_agente') == id_usuario]
    return render_template('resolverTicket.html', tickets_pendientes=pendientes)


@app.route('/ticket/<int:id_ticket>')
@login_required
def ver_ticket(id_ticket):
    ticket = obtenerTicket(id_ticket)
    if not ticket:
        abort(404)
    detalles = obtenerDetallesTicket(id_ticket)
    return render_template('verTicket.html', ticket=ticket,
                           detalles=detalles)


@app.route('/ticket/<int:id_ticket>/resolver')
@login_required
def form_resolver_ticket(id_ticket):
    ticket   = obtenerTicket(id_ticket)
    if not ticket or ticket['estado'] != 'en_progreso':
        return redirect(url_for('listar_tickets'))
    if ticket.get('id_agente') != session['usuario']['id_usuario']:
        return redirect(url_for('listar_tickets'))
    nivel_usuario = session['usuario'].get('nivel', 1)
    agentes_disponibles = listarUsuariosNivelMenor(nivel_usuario) if nivel_usuario > 1 else []
    detalles = obtenerDetallesTicket(id_ticket)
    return render_template('resolverTicket.html', ticket=ticket,
                           nivel_usuario=nivel_usuario,
                           agentes_disponibles=agentes_disponibles,
                           detalles=detalles)


@app.route('/ticket/<int:id_ticket>/calificar', methods=['GET'])
@login_required
def form_calificar_ticket(id_ticket):
    ticket = obtenerTicket(id_ticket)
    if not ticket:
        abort(404)
    if ticket['estado'] != 'resuelto':
        abort(400)
    if ticket['id_solicitante'] != session['usuario']['id_usuario']:
        return redirect(url_for('listar_tickets'))
    detalles = obtenerDetallesTicket(id_ticket)
    calificacion_actual = next((d for d in detalles if d.get('activo') and d.get('calificacion_estrellas')), None)
    return render_template('calificarTicket.html',
                           ticket=ticket,
                           calificacion_actual=calificacion_actual,
                           mensaje=None,
                           tipo=None)


@app.route('/ticket/<int:id_ticket>/editar', methods=['GET'])
@login_required
def form_editar_ticket(id_ticket):
    ticket = obtenerTicket(id_ticket)
    if not ticket or ticket['estado'] != 'solicitado':
        return redirect(url_for('listar_tickets'))
    if ticket['id_solicitante'] != session['usuario']['id_usuario']:
        return redirect(url_for('listar_tickets'))
    aplicaciones = listarAplicaciones()
    return render_template('editarTicket.html', ticket=ticket, aplicaciones=aplicaciones)


# ── ASIGNAR AGENTE ────────────────────────────────────────────────────────────
@app.route('/ticket/<int:id_ticket>/asignar')
@login_required
def form_asignar_ticket(id_ticket):
    ticket = obtenerTicket(id_ticket)
    if not ticket or ticket['estado'] != 'solicitado':
        return redirect(url_for('listar_tickets'))
    agentes = listarPosiblesAgentes()
    aplicaciones = listarAplicaciones()
    return render_template('asignarTicket.html', ticket=ticket, agentes=agentes, aplicaciones=aplicaciones)


# ──────────────────────────────────────────────────────────────
#  APLICACIONES CRUD  (solo admin)
# ──────────────────────────────────────────────────────────────

@app.route('/aplicacion/nueva', methods=['GET'])
@login_required
def form_nueva_aplicacion():
    return render_template('formularioAplicacion.html', aplicacion=None)


@app.route('/aplicacion/<int:id_aplicacion>/editar', methods=['GET'])
@login_required
def form_editar_aplicacion(id_aplicacion):
    app = obtenerAplicacion(id_aplicacion)
    if not app:
        abort(404)
    return render_template('formularioAplicacion.html', aplicacion=app)


# ──────────────────────────────────────────────────────────────
#  PROYECTOS DE SOFTWARE
# ──────────────────────────────────────────────────────────────

@app.route('/proyecto/nuevo')
@login_required
def form_proyecto():
    responsables = obtenerUsuarios()
    return render_template('nuevoProyecto.html', responsables=responsables)


@app.route('/proyecto/<int:id_proyecto>/editar')
@login_required
def form_editar_proyecto(id_proyecto):
    proyecto = obtenerProyecto(id_proyecto)
    if not proyecto:
        abort(404)
    responsables = obtenerUsuarios()
    from proyectoAD import obtenerResponsablesProyecto
    ids_responsables_actuales = obtenerResponsablesProyecto(id_proyecto)
    return render_template('editarProyecto.html',
                           proyecto=proyecto,
                           responsables=responsables,
                           ids_responsables_actuales=ids_responsables_actuales)


@app.route('/proyectos')
@login_required
def listar_proyectos():
    try:
        proyectos = listarProyectos()
    except Exception as e:
        print(f'Error al listar proyectos: {e}')
        abort(500)
    return render_template('listaProyectos.html', proyectos=proyectos)

@app.route('/api/proyecto/<int:id_proyecto>/eliminar', methods=['POST'])
@jwt_required()
@login_required
def api_eliminar_proyecto(id_proyecto):
    if tieneActividadesPendientes(id_proyecto):
        return jsonify({'ok': False, 'mensaje': 'No se puede eliminar: el proyecto tiene actividades pendientes.'})
    ok = eliminarProyecto(id_proyecto)
    if ok:
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'mensaje': 'No se pudo eliminar el proyecto.'})


@app.route('/api/reporte/proyecto/<int:id_proyecto>/actividades')
@jwt_required()
@login_required
def api_reporte_actividades(id_proyecto):
    actividades = reporteActividadesPorProyecto(id_proyecto)
    resumen     = resumenActividadesReporte(id_proyecto)
    return jsonify({'ok': True, 'actividades': actividades, 'resumen': resumen})


@app.route('/proyecto/<int:id_proyecto>/gestion')
@login_required
def gestion_proyecto(id_proyecto):
    proyecto = obtenerProyecto(id_proyecto)
    if not proyecto:
        abort(404)

    actividades     = listarActividades(id_proyecto)
    todos_proyectos = listarProyectos()

    fecha_fin = proyecto['fecha_fin_plan']
    from datetime import date
    hoy       = date.today()
    en_riesgo = (fecha_fin < hoy) and (proyecto['estado'] != 'completado')

    return render_template(
        'gestionProyecto.html',
        proyecto        = proyecto,
        actividades     = actividades,
        todos_proyectos = todos_proyectos,
        en_riesgo       = en_riesgo,
    )


@app.route('/proyecto/<int:id_proyecto>/avances')
@login_required
def historial_avances(id_proyecto):
    proyecto = obtenerProyecto(id_proyecto)
    if not proyecto:
        abort(404)
    avances = listarAvances(id_proyecto)

    # Datos cronológicos para el gráfico (listas Python puras, sin json.dumps)
    avances_cronologicos = list(reversed(avances))
    fechas_grafico      = [av['fecha_reporte'].strftime('%d/%m') for av in avances_cronologicos] if avances else []
    porcentajes_grafico = [float(av['porcentaje_avance']) for av in avances_cronologicos]         if avances else []

    return render_template('avancesProyecto.html',
                           proyecto=proyecto,
                           avances=avances,
                           fechas_grafico=fechas_grafico,
                           porcentajes_grafico=porcentajes_grafico)


@app.route('/proyecto/<int:id_proyecto>/avances/nuevo', methods=['GET'])
@login_required
def form_nuevo_avance(id_proyecto):
    proyecto = obtenerProyecto(id_proyecto)
    if not proyecto:
        abort(404)

    resumen     = resumenActividadesPorProyecto()
    fila_actual = next((r for r in resumen
                        if (r['id_proyecto'] if isinstance(r, dict) else r[0]) == id_proyecto), None)

    if fila_actual:
        if isinstance(fila_actual, dict):
            valor_raw = fila_actual.get('porcentaje_avance_real')
        else:
            valor_raw = fila_actual[3]
        pct_calculado = round(float(valor_raw)) if valor_raw is not None else 0
    else:
        pct_calculado = 0

    if proyecto.get('estado') == 'completado':
        pct_calculado = 100

    from datetime import date
    hoy_mostrar = date.today().strftime('%d/%m/%Y')

    usuarios = obtenerUsuarios()
    return render_template('nuevoAvance.html',
                           proyecto=proyecto,
                           hoy_mostrar=hoy_mostrar,
                           pct_calculado=pct_calculado,
                           usuarios=usuarios)


# ──────────────────────────────────────────────────────────────
#  INDICADORES DE SOPORTE
# ──────────────────────────────────────────────────────────────

@app.route('/indicadores')
@login_required
def indicadores_soporte():
    mes_inicio  = request.args.get('mes_inicio', '')  or None
    anio_inicio = request.args.get('anio_inicio', '') or None
    mes_fin     = request.args.get('mes_fin', '')     or None
    anio_fin    = request.args.get('anio_fin', '')    or None

    resumen           = resumenKPI()
    por_app           = kpiPorAplicacion()
    por_prioridad     = kpiPorPrioridad()
    por_agente        = kpiPorAgente()
    por_mes           = kpiPorMes()
    sprint_activo     = kpiSprintsActivos()
    satisfaccion      = kpiSatisfaccion()
    comentarios       = comentariosCalificacionesRecientes()
    proyectos_estado  = kpiProyectosPorEstado()
    velocity_sprints  = kpiVelocityPorSprint()
    carga_programador = kpiCargaPorProgramador()
    tiempo_respuesta  = kpiTiempoRespuesta()
    por_intensidad    = kpiPorIntensidad()
    sla_por_agente    = kpiSLAPorAgente()
    salud_proyectos   = kpiProyectosPorSalud()
    avance_proyectos  = kpiAvancePromedioPorProyecto()
    tiempo_por_app    = kpiTiempoResolucionPorAplicacion()
    cancelados        = kpiCancelados()
    ranking_apps      = kpiRankingAppsProblemáticas()
    por_tipo          = kpiPorTipo()
    top5_lentos       = kpiTop5MasLentos()
    act_por_estado    = kpiActividadesPorEstado()
    sin_carga         = kpiProgramadoresSinCarga()
    proy_vencidos     = kpiProyectosVencidos()

    return render_template('indicadores.html',
                           resumen=resumen,
                           por_app=por_app,
                           por_prioridad=por_prioridad,
                           por_agente=por_agente,
                           por_mes=por_mes,
                           sprint_activo=sprint_activo,
                           satisfaccion=satisfaccion,
                           comentarios=comentarios,
                           proyectos_estado=proyectos_estado,
                           velocity_sprints=velocity_sprints,
                           carga_programador=carga_programador,
                           tiempo_respuesta=tiempo_respuesta,
                           por_intensidad=por_intensidad,
                           sla_por_agente=sla_por_agente,
                           salud_proyectos=salud_proyectos,
                           avance_proyectos=avance_proyectos,
                           tiempo_por_app=tiempo_por_app,
                           cancelados=cancelados,
                           ranking_apps=ranking_apps,
                           por_tipo=por_tipo,
                           top5_lentos=top5_lentos,
                           act_por_estado=act_por_estado,
                           sin_carga=sin_carga,
                           proy_vencidos=proy_vencidos,
                           mes_inicio=mes_inicio   or '',
                           anio_inicio=anio_inicio or '',
                           mes_fin=mes_fin         or '',
                           anio_fin=anio_fin       or '')


# ──────────────────────────────────────────────────────────────
#  REPORTES
# ──────────────────────────────────────────────────────────────

def limpiar(valor):
    if valor is None:
        return None
    valor = valor.strip()
    return valor if valor else None


@app.route('/reportes')
@login_required
def gestion_reportes():
    fecha_inicio  = limpiar(request.args.get('fecha_inicio', ''))
    fecha_fin     = limpiar(request.args.get('fecha_fin', ''))
    id_aplicacion = limpiar(request.args.get('id_aplicacion', ''))
    estado        = limpiar(request.args.get('estado', ''))
    prioridad     = limpiar(request.args.get('prioridad', ''))

    resumen             = reporteResumen()
    tickets_app         = reporteTicketsPorApp()
    tickets_tipo        = reporteTicketsPorTipo()
    story_points        = reporteStoryPointsPorProgramador()
    carryover           = reporteCarryoverPorProgramador()
    aplicaciones        = obtenerAplicaciones()
    tickets_filtrados   = reporteTicketsFiltrados(fecha_inicio, fecha_fin,
                                                  id_aplicacion, estado, prioridad)
    proyectos_estado    = reporteProyectosPorEstado()
    proyectos_riesgo    = reporteProyectosEnRiesgo()
    rendimiento_sprint  = reporteRendimientoPorSprint()
    proyectos_filtrados = reporteProyectosFiltrados()
    responsables        = obtenerResponsables()
    sla_por_app         = reporteSLAPorAplicacion()
    tickets_por_estado  = reporteTicketsPorEstado()
    agentes_metricas    = reporteAgentesMetricas()
    tendencia_mes       = reporteTendenciaPorMes()

    return render_template('reportes.html',
                           resumen=resumen,
                           tickets_app=tickets_app,
                           tickets_tipo=tickets_tipo,
                           story_points=story_points,
                           carryover=carryover,
                           aplicaciones=aplicaciones,
                           tickets_filtrados=tickets_filtrados,
                           proyectos_estado=proyectos_estado,
                           proyectos_riesgo=proyectos_riesgo,
                           rendimiento_sprint=rendimiento_sprint,
                           proyectos_filtrados=proyectos_filtrados,
                           responsables=responsables,
                           sla_por_app=sla_por_app,
                           tickets_por_estado=tickets_por_estado,
                           agentes_metricas=agentes_metricas,
                           tendencia_mes=tendencia_mes,
                           fecha_inicio=fecha_inicio   or '',
                           fecha_fin=fecha_fin         or '',
                           id_aplicacion=id_aplicacion or '',
                           estado=estado               or '',
                           prioridad=prioridad         or '')


# ──────────────────────────────────────────────────────────────
#  EXPORTAR EXCEL
# ──────────────────────────────────────────────────────────────

def _estilo_cabecera(ws, fila, columnas, color_hex="1a5276"):
    """Aplica estilo de cabecera a una fila del worksheet."""
    fill   = PatternFill("solid", fgColor=color_hex)
    fuente = Font(bold=True, color="FFFFFF", size=10)
    borde  = Border(
        bottom=Side(style='medium', color='FFFFFF'),
        right =Side(style='thin',   color='FFFFFF')
    )
    for col_idx, titulo in enumerate(columnas, 1):
        cell = ws.cell(row=fila, column=col_idx, value=titulo)
        cell.fill      = fill
        cell.font      = fuente
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = borde

def _autowidth(ws, extra=4):
    """Ajusta el ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + extra, 50)

def _fila_alterna(ws, fila, n_cols, par=True):
    color = "EBF5FB" if par else "FFFFFF"
    fill  = PatternFill("solid", fgColor=color)
    for c in range(1, n_cols + 1):
        ws.cell(row=fila, column=c).fill = fill

@app.route('/reportes/exportar-excel')
@login_required
def exportar_excel():
    fecha_inicio  = limpiar(request.args.get('fecha_inicio', ''))
    fecha_fin     = limpiar(request.args.get('fecha_fin', ''))
    id_aplicacion = limpiar(request.args.get('id_aplicacion', ''))
    estado        = limpiar(request.args.get('estado', ''))
    prioridad     = limpiar(request.args.get('prioridad', ''))

    tickets        = reporteTicketsFiltrados(fecha_inicio, fecha_fin, id_aplicacion, estado, prioridad)
    resumen        = reporteResumen()
    tickets_app    = reporteTicketsPorApp()
    sprints        = reporteRendimientoPorSprint()
    story_pts      = reporteStoryPointsPorProgramador()
    carryover      = reporteCarryoverPorProgramador()
    proyectos      = reporteProyectosFiltrados()
    tiempo_res     = kpiPorAgente()          # promedio horas resolución por agente

    wb = Workbook()

    # ── HOJA 1: Tickets filtrados ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Tickets"
    ws1.row_dimensions[1].height = 30

    cabeceras_t = [
        "#", "Título", "Aplicación", "Tipo", "Prioridad", "Intensidad",
        "Estado", "Solicitante", "Agente", "Fecha Apertura", "Fecha Asignación",
        "Fecha Solución", "Fecha Cierre", "SLA (h)", "Tiempo resolución",
        "SLA cumplido", "Calificación", "Observación"
    ]
    _estilo_cabecera(ws1, 1, cabeceras_t, "1a5276")

    for i, t in enumerate(tickets, 2):
        mins = t.get('minutos_resolucion')
        if mins is not None:
            horas, resto = divmod(int(mins), 60)
            tiempo_str   = f"{horas}h {resto}m"
        else:
            tiempo_str = "—"

        fila = [
            t.get('id_ticket'),
            t.get('titulo'),
            t.get('aplicacion'),
            t.get('tipo', '').capitalize(),
            t.get('prioridad', '').capitalize(),
            t.get('intensidad', '').capitalize(),
            (t.get('estado') or '').replace('_', ' ').capitalize(),
            t.get('solicitante'),
            t.get('agente') or '—',
            t.get('fecha_apertura') or '—',
            t.get('fecha_asignacion') or '—',
            t.get('fecha_solucion') or '—',
            t.get('fecha_cierre') or '—',
            t.get('sla_horas') or '—',
            tiempo_str,
            t.get('sla_cumplido') or '—',
            t.get('calificacion') or '—',
            t.get('obs_calificacion') or '—',
        ]
        for col_idx, valor in enumerate(fila, 1):
            cell = ws1.cell(row=i, column=col_idx, value=valor)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            # Color SLA
            if col_idx == 16:
                if valor == 'SI':
                    cell.font = Font(color="1E8449", bold=True)
                elif valor == 'NO':
                    cell.font = Font(color="C0392B", bold=True)
        _fila_alterna(ws1, i, len(cabeceras_t), i % 2 == 0)
        ws1.row_dimensions[i].height = 18
    _autowidth(ws1)
    ws1.freeze_panes = "A2"

    # ── HOJA 2: Tickets por Aplicación ────────────────────────
    ws2 = wb.create_sheet("Por Aplicación")
    _estilo_cabecera(ws2, 1, ["Aplicación", "Total", "Pendientes", "Cerrados"], "1f618d")
    for i, r in enumerate(tickets_app, 2):
        r = dict(r)
        ws2.cell(row=i, column=1, value=r.get('aplicacion'))
        ws2.cell(row=i, column=2, value=r.get('total'))
        ws2.cell(row=i, column=3, value=r.get('pendientes') or 0)
        ws2.cell(row=i, column=4, value=r.get('cerrados') or 0)
        _fila_alterna(ws2, i, 4, i % 2 == 0)
    _autowidth(ws2)
    ws2.freeze_panes = "A2"

    # ── HOJA 3: Rendimiento por Sprint ────────────────────────
    ws3 = wb.create_sheet("Sprints")
    _estilo_cabecera(ws3, 1,
        ["Sprint", "Proyecto", "Capacidad (pts)", "Completados (pts)", "Pendientes (pts)",
         "% Completado", "Estado"], "1b4f72")
    for i, s in enumerate(sprints, 2):
        s = dict(s)
        cap  = s.get('capacidad_pts') or 0
        comp = s.get('pts_completados') or 0
        pct  = round(comp / cap * 100, 1) if cap else 0
        ws3.cell(row=i, column=1, value=s.get('sprint'))
        ws3.cell(row=i, column=2, value=s.get('proyecto'))
        ws3.cell(row=i, column=3, value=cap)
        ws3.cell(row=i, column=4, value=comp)
        ws3.cell(row=i, column=5, value=s.get('pts_pendientes') or 0)
        ws3.cell(row=i, column=6, value=pct)
        ws3.cell(row=i, column=7, value=(s.get('estado_sprint') or '').capitalize())
        _fila_alterna(ws3, i, 7, i % 2 == 0)
    _autowidth(ws3)
    ws3.freeze_panes = "A2"

    # ── HOJA 4: Story Points por Programador ──────────────────
    ws4 = wb.create_sheet("Story Points")
    _estilo_cabecera(ws4, 1,
        ["Programador", "Pts Completados", "Pts Asignados", "% Completado"], "145a32")
    for i, r in enumerate(story_pts, 2):
        r = dict(r)
        asig = r.get('pts_asignados') or 0
        comp = r.get('pts_completados') or 0
        pct  = round(comp / asig * 100, 1) if asig else 0
        ws4.cell(row=i, column=1, value=r.get('programador'))
        ws4.cell(row=i, column=2, value=comp)
        ws4.cell(row=i, column=3, value=asig)
        ws4.cell(row=i, column=4, value=pct)
        _fila_alterna(ws4, i, 4, i % 2 == 0)
    _autowidth(ws4)
    ws4.freeze_panes = "A2"

    # ── HOJA 5: Carryover ─────────────────────────────────────
    ws5 = wb.create_sheet("Carryover")
    _estilo_cabecera(ws5, 1,
        ["Programador", "Sprints con Carryover", "Pts Carryover"], "6e2f1a")
    for i, r in enumerate(carryover, 2):
        r = dict(r)
        ws5.cell(row=i, column=1, value=r.get('programador'))
        ws5.cell(row=i, column=2, value=r.get('sprints_con_carryover'))
        ws5.cell(row=i, column=3, value=r.get('pts_carryover'))
        _fila_alterna(ws5, i, 3, i % 2 == 0)
    _autowidth(ws5)
    ws5.freeze_panes = "A2"

    # ── HOJA 6: Proyectos ─────────────────────────────────────
    ws6 = wb.create_sheet("Proyectos")
    _estilo_cabecera(ws6, 1,
        ["#", "Proyecto", "Responsable", "Estado", "Inicio", "Fecha Fin",
         "Días restantes", "Salud"], "1b2631")
    for i, p in enumerate(proyectos, 2):
        p = dict(p)
        salud_map = {'completado': 'Completado', 'vencido': 'Vencido',
                     'por_vencer': 'Por vencer', 'ok': 'OK'}
        ws6.cell(row=i, column=1, value=p.get('id_proyecto'))
        ws6.cell(row=i, column=2, value=p.get('nombre'))
        ws6.cell(row=i, column=3, value=p.get('responsable'))
        ws6.cell(row=i, column=4, value=(p.get('estado') or '').replace('_', ' ').capitalize())
        ws6.cell(row=i, column=5, value=str(p.get('fecha_inicio') or '—'))
        ws6.cell(row=i, column=6, value=str(p.get('fecha_fin_plan') or '—'))
        ws6.cell(row=i, column=7, value=p.get('dias_restantes'))
        ws6.cell(row=i, column=8, value=salud_map.get(p.get('salud'), '—'))
        # Color en columna Salud
        salud_cell = ws6.cell(row=i, column=8)
        if p.get('salud') == 'vencido':
            salud_cell.font = Font(color="C0392B", bold=True)
        elif p.get('salud') == 'por_vencer':
            salud_cell.font = Font(color="D68910", bold=True)
        elif p.get('salud') == 'ok':
            salud_cell.font = Font(color="1E8449", bold=True)
        _fila_alterna(ws6, i, 8, i % 2 == 0)
    _autowidth(ws6)
    ws6.freeze_panes = "A2"

    # ── HOJA 7: Tiempo de resolución por agente ───────────────
    ws7 = wb.create_sheet("Tiempo Resolución")
    _estilo_cabecera(ws7, 1,
        ["Agente", "Tickets atendidos", "Resueltos",
         "Promedio resolución (h)", "Cumplimiento SLA (%)"], "4a235a")
    for i, a in enumerate(tiempo_res, 2):
        a = dict(a)
        # Buscar pct_sla del agente en sla_por_agente si existe
        ws7.cell(row=i, column=1, value=a.get('agente'))
        ws7.cell(row=i, column=2, value=a.get('total_atendidos'))
        ws7.cell(row=i, column=3, value=a.get('resueltos') or 0)
        ws7.cell(row=i, column=4, value=a.get('promedio_horas') or '—')
        _fila_alterna(ws7, i, 5, i % 2 == 0)
    _autowidth(ws7)
    ws7.freeze_panes = "A2"

    # ── Enviar el archivo ──────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.read())
    response.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_agrovision.xlsx'
    return response


@app.route('/reportes/exportar-excel-proyectos')
@login_required
def exportar_excel_proyectos():
    estado        = limpiar(request.args.get('estado', ''))
    id_responsable = limpiar(request.args.get('id_responsable', ''))

    proyectos    = reporteProyectosFiltrados(estado or None, id_responsable or None)
    sprints      = reporteRendimientoPorSprint()

    wb = Workbook()

    # ── HOJA 1: Proyectos ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Proyectos"
    ws1.row_dimensions[1].height = 30
    _estilo_cabecera(ws1, 1,
        ["#", "Proyecto", "Responsable", "Estado", "Inicio", "Fecha Fin",
         "Días restantes", "Salud"], "1b2631")
    for i, p in enumerate(proyectos, 2):
        p = dict(p)
        salud_map = {'completado': 'Completado', 'vencido': 'Vencido',
                     'por_vencer': 'Por vencer', 'ok': 'OK'}
        ws1.cell(row=i, column=1, value=p.get('id_proyecto'))
        ws1.cell(row=i, column=2, value=p.get('nombre'))
        ws1.cell(row=i, column=3, value=p.get('responsable'))
        ws1.cell(row=i, column=4, value=(p.get('estado') or '').replace('_', ' ').capitalize())
        ws1.cell(row=i, column=5, value=str(p.get('fecha_inicio') or '—'))
        ws1.cell(row=i, column=6, value=str(p.get('fecha_fin_plan') or '—'))
        ws1.cell(row=i, column=7, value=p.get('dias_restantes'))
        ws1.cell(row=i, column=8, value=salud_map.get(p.get('salud'), '—'))
        salud_cell = ws1.cell(row=i, column=8)
        if p.get('salud') == 'vencido':
            salud_cell.font = Font(color="C0392B", bold=True)
        elif p.get('salud') == 'por_vencer':
            salud_cell.font = Font(color="D68910", bold=True)
        elif p.get('salud') == 'ok':
            salud_cell.font = Font(color="1E8449", bold=True)
        _fila_alterna(ws1, i, 8, i % 2 == 0)
        ws1.row_dimensions[i].height = 18
    _autowidth(ws1)
    ws1.freeze_panes = "A2"

    # ── HOJA 2: Rendimiento por Sprint ────────────────────────
    ws2 = wb.create_sheet("Sprints")
    _estilo_cabecera(ws2, 1,
        ["Sprint", "Proyecto", "Capacidad (pts)", "Completados (pts)",
         "Pendientes (pts)", "% Completado", "Estado"], "1b4f72")
    for i, s in enumerate(sprints, 2):
        s = dict(s)
        cap  = s.get('capacidad_pts') or 0
        comp = s.get('pts_completados') or 0
        pct  = round(comp / cap * 100, 1) if cap else 0
        ws2.cell(row=i, column=1, value=s.get('sprint'))
        ws2.cell(row=i, column=2, value=s.get('proyecto'))
        ws2.cell(row=i, column=3, value=cap)
        ws2.cell(row=i, column=4, value=comp)
        ws2.cell(row=i, column=5, value=s.get('pts_pendientes') or 0)
        ws2.cell(row=i, column=6, value=pct)
        ws2.cell(row=i, column=7, value=(s.get('estado_sprint') or '').capitalize())
        _fila_alterna(ws2, i, 7, i % 2 == 0)
    _autowidth(ws2)
    ws2.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.read())
    response.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_proyectos.xlsx'
    return response


# ──────────────────────────────────────────────────────────────
#  ACTIVIDADES
# ──────────────────────────────────────────────────────────────

@app.route('/actividad/nueva')
@login_required
def form_actividad():
    id_proyecto_pre = request.args.get('id_proyecto', type=int)
    proyectos       = listarProyectos()
    sprints         = listarTodosSprints()
    proximo_codigo  = proximoCodigo()

    # Dict {id_proyecto: [lista de asignados]} — se pasa como dict Python normal
    asignados_json = {}
    for p in proyectos:
        pid   = p['id_proyecto']
        lista = listarAsignadosPorProyecto(pid)
        asignados_json[pid] = [{'id_usuario': u['id_usuario'],
                                'nombre_completo': u['nombre_completo']}
                               for u in lista]

    return render_template('nuevaActividad.html',
                           proyectos=proyectos,
                           sprints=sprints,
                           proximo_codigo=proximo_codigo,
                           asignados_json=asignados_json,
                           id_proyecto_pre=id_proyecto_pre)


@app.route('/actividad/<int:id_actividad>/editar')
@login_required
def form_editar_actividad(id_actividad):
    from actividadAD import obtenerActividad
    actividad = obtenerActividad(id_actividad)
    if not actividad:
        abort(404)
    proyectos  = listarProyectos()
    sprints    = listarTodosSprints()
    asignados_json = {}
    for p in proyectos:
        pid   = p['id_proyecto']
        lista = listarAsignadosPorProyecto(pid)
        asignados_json[pid] = [{'id_usuario': u['id_usuario'],
                                'nombre_completo': u['nombre_completo']}
                               for u in lista]
    return render_template('editarActividad.html',
                           actividad=actividad,
                           proyectos=proyectos,
                           sprints=sprints,
                           asignados_json=asignados_json)


# ──────────────────────────────────────────────────────────────
#  USUARIOS · Lista, Perfil e Historial
# ──────────────────────────────────────────────────────────────

@app.route('/usuarios')
@login_required
def lista_usuarios():
    nombre   = request.args.get('nombre', '').strip()
    try:
        usuarios = listarUsuariosCompleto()
    except Exception as e:
        print(f'Error al listar usuarios: {e}')
        abort(500)
    if nombre:
        nombre_lower = nombre.lower()
        usuarios = [u for u in usuarios
                    if nombre_lower in u['nombre_completo'].lower()
                    or (u['apellido'] and nombre_lower in u['apellido'].lower())]
    return render_template('listaUsuarios.html', usuarios=usuarios, nombre_busqueda=nombre)


@app.route('/usuario/<int:id_usuario>/perfil')
@login_required
def perfil_usuario(id_usuario):
    perfil = obtenerPerfilUsuario(id_usuario)
    if not perfil:
        abort(404)
    stats = estadisticasUsuario(id_usuario)

    tickets_json   = [dict(t) for t in stats['tickets_por_tipo']]
    proyectos_json = [dict(p) for p in stats['proyectos_por_estado']]

    return render_template('perfilUsuario.html',
                           perfil=perfil,
                           stats=stats,
                           tickets_json=tickets_json,
                           proyectos_json=proyectos_json)


@app.route('/usuario/<int:id_usuario>/historial')
@login_required
def historial_usuario(id_usuario):
    perfil = obtenerPerfilUsuario(id_usuario)
    if not perfil:
        abort(404)
    items   = historialParticipacionUsuario(id_usuario)
    resumen = resumenHistorialUsuario(id_usuario)
    return render_template('historialUsuario.html',
                           perfil=perfil,
                           items=items,
                           resumen=resumen)


# ──────────────────────────────────────────────────────────────
#  API JSON  (jsonify)
# ──────────────────────────────────────────────────────────────

@app.route('/api/login', methods=['POST'])
def api_login():
    if request.is_json:
        correo   = (request.json.get('correo') or '').strip()
        password = (request.json.get('password') or '')
    else:
        correo   = request.form.get('correo', '').strip()
        password = request.form.get('password', '')
    try:
        ok, mensaje, datos = autenticarUsuario(correo, password)
    except Exception as e:
        print(f'Error de conexión al autenticar: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error de conexión con la base de datos.',
                        'error_servidor': True}), 500
    if ok:
        session['usuario'] = datos
        session.permanent = True
        token = jwt.encode({'usuario': datos}, JWT_SECRET, algorithm=JWT_ALGORITHM)
        if isinstance(token, bytes):
            token = token.decode('utf-8')
        return jsonify({'ok': True, 'usuario': datos, 'token': token})
    return jsonify({'ok': False, 'mensaje': mensaje})


@app.route('/api/usuario/me')
@jwt_required()
@login_required
def api_usuario_me():
    usuario_session = session.get('usuario')
    id_usuario = usuario_session.get('id_usuario') if usuario_session else None
    if not id_usuario:
        return jsonify({'ok': False, 'mensaje': 'Sesión inválida'}), 401
    usuario = obtenerPerfilUsuario(id_usuario)
    if not usuario or not usuario.get('activo'):
        return jsonify({'ok': False, 'mensaje': 'Usuario no encontrado o inactivo'}), 404
    permisos = []
    if usuario.get('id_rol'):
        permisos = obtenerPermisosPorRol(usuario['id_rol'])
    return jsonify({
        'ok': True,
        'usuario': {
            'id_usuario'     : usuario['id_usuario'],
            'nombre_completo': usuario['nombre_completo'],
            'correo'         : usuario['correo'],
            'rol_nombre'     : usuario.get('rol_nombre') or '',
            'rol_id'         : usuario.get('id_rol'),
            'nivel'          : usuario.get('nivel') or 1,
            'permisos'       : permisos,
        }
    })


@app.route('/api/indicadores')
@jwt_required()
@login_required
def api_indicadores():
    return jsonify({
        'por_app':       kpiPorAplicacion(),
        'por_prioridad': kpiPorPrioridad(),
        'por_mes':       kpiPorMes(),
    })


@app.route('/api/tickets')
@jwt_required()
@login_required
def api_tickets():
    estado = request.args.get('estado', '') or None
    texto  = request.args.get('texto',  '') or None
    todos  = listarTickets()
    if estado:
        todos = [t for t in todos if t['estado'] == estado]
    if texto:
        texto = texto.lower()
        todos = [t for t in todos if texto in (t['titulo'] or '').lower()
                                  or texto in (t['nombre_aplicacion'] or '').lower()
                                  or texto in (t['nombre_solicitante'] or '').lower()]
    # Convertir fechas a string para que jsonify las serialice
    for t in todos:
        for k, v in t.items():
            if hasattr(v, 'isoformat'):
                t[k] = v.isoformat()
    return jsonify(todos)


@app.route('/api/actividad/<int:id_actividad>/estado', methods=['POST'])
@jwt_required()
@login_required
def api_estado_actividad(id_actividad):
    data         = request.get_json()
    nuevo_estado = data.get('estado') if data else None
    if not nuevo_estado:
        return jsonify({'ok': False, 'mensaje': 'Estado no enviado'}), 400
    actualizarEstadoActividad(id_actividad, nuevo_estado)
    return jsonify({'ok': True, 'estado': nuevo_estado})


@app.route('/api/actividad/<int:id_actividad>/desbloquear', methods=['POST'])
@jwt_required()
@login_required
def api_desbloquear_actividad(id_actividad):
    estado_retorno = desbloquearActividad(id_actividad)
    return jsonify({'ok': True, 'estado': estado_retorno})


@app.route('/api/actividad/<int:id_actividad>/eliminar', methods=['POST'])
@jwt_required()
@login_required
def api_eliminar_actividad(id_actividad):
    ok = eliminarActividad(id_actividad)
    if ok:
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'mensaje': 'Solo se pueden eliminar actividades canceladas.'})


@app.route('/api/actividades')
@jwt_required()
@login_required
def api_listar_actividades():
    id_proyecto = request.args.get('id_proyecto', type=int) or None
    return jsonify({'ok': True, 'actividades': listarActividades(id_proyecto)})


@app.route('/api/proyecto/<int:id_proyecto>/porcentaje')
@jwt_required()
@login_required
def api_porcentaje_proyecto(id_proyecto):
    resumen     = resumenActividadesPorProyecto()
    fila_actual = next((r for r in resumen
                        if (r['id_proyecto'] if isinstance(r, dict) else r[0]) == id_proyecto), None)
    if fila_actual:
        if isinstance(fila_actual, dict):
            valor_raw = fila_actual.get('porcentaje_avance_real')
        else:
            valor_raw = fila_actual[3]
        pct = round(float(valor_raw)) if valor_raw is not None else 0
    else:
        pct = 0

    proyecto = obtenerProyecto(id_proyecto)
    if proyecto and proyecto.get('estado') == 'completado':
        pct = 100

    return jsonify({'porcentaje': pct})

@app.route('/api/proyecto/<int:id_proyecto>/avances-grafico')
@jwt_required()
@login_required
def api_avances_grafico(id_proyecto):
    proyecto = obtenerProyecto(id_proyecto)
    avances = listarAvances(id_proyecto)
    avances_cronologicos = list(reversed(avances))
    
    fechas = []
    porcentajes_reales = []
    porcentajes_estimados = []

    if proyecto and proyecto.get('fecha_inicio') and proyecto.get('fecha_fin_plan'):
        inicio = proyecto['fecha_inicio']
        fin = proyecto['fecha_fin_plan']
        # Calculamos la duración total del proyecto en días
        duracion_total = (fin - inicio).days

        for av in avances_cronologicos:
            fecha_rep = av['fecha_reporte']
            fechas.append(fecha_rep.strftime('%d/%m'))
            porcentajes_reales.append(float(av['porcentaje_avance']))
            
            # Cálculo del Avance Estimado (Lineal / Roadmap)
            if duracion_total > 0:
                dias_transcurridos = (fecha_rep - inicio).days
                esperado = (dias_transcurridos / duracion_total) * 100
                # Evitamos que baje de 0% o pase de 100%
                esperado = max(0, min(100, esperado))
            else:
                esperado = 100
                
            porcentajes_estimados.append(round(esperado, 1))
    else:
        # Fallback de seguridad si el proyecto no tiene fechas
        fechas = [av['fecha_reporte'].strftime('%d/%m') for av in avances_cronologicos]
        porcentajes_reales = [float(av['porcentaje_avance']) for av in avances_cronologicos]
        porcentajes_estimados = [0] * len(avances_cronologicos)

    return jsonify({
        'fechas': fechas,
        'porcentajes': porcentajes_reales,
        'estimados': porcentajes_estimados
    })

@app.route('/api/avance/<int:id_avance>/eliminar', methods=['POST'])
@jwt_required()
@login_required
def api_eliminar_avance(id_avance):
    try:
        eliminarAvance(id_avance)
        return jsonify({'ok': True})
    except Exception as e:
        print(f"Error de integridad al eliminar avance: {e}")
        return jsonify({'ok': False, 'mensaje': 'No se puede eliminar por restricciones de integridad en la base de datos.'})


# ──────────────────────────────────────────────────────────────
#  API TICKETS CRUD
# ──────────────────────────────────────────────────────────────

@app.route('/api/ticket/guardar', methods=['POST'])
@jwt_required()
@login_required
def api_guardar_ticket():
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    usuario_session = session.get('usuario')
    id_solicitante = usuario_session.get('id_usuario') if usuario_session else None
    if not id_solicitante:
        return jsonify({'ok': False, 'mensaje': 'Sesión inválida'}), 401
    try:
        obj = Ticket(
            titulo                = data['titulo'],
            tipo                  = data['tipo'],
            id_solicitante        = id_solicitante,
            id_aplicacion         = data['id_aplicacion'],
            descripcion           = data['descripcion'],
            link_img_descripcion  = (data.get('link_img_descripcion') or '').strip() or None,
        )
        insertarTicket(obj)
        return jsonify({'ok': True})
    except KeyError as e:
        return jsonify({'ok': False, 'mensaje': f'Campo requerido: {e}'}), 400
    except Exception as e:
        print(f'Error al guardar ticket: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error al guardar ticket.'}), 500


@app.route('/api/ticket/<int:id_ticket>/resolver', methods=['POST'])
@jwt_required()
@login_required
def api_resolver_ticket(id_ticket):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    id_agente           = session['usuario']['id_usuario']
    notas               = (data.get('notas_resolucion') or '').strip()
    link_img_resolucion = (data.get('link_img_resolucion') or '').strip() or None
    resolverTicket(id_ticket, id_agente, 'resuelto', notas, link_img_resolucion)
    return jsonify({'ok': True})


@app.route('/api/ticket/<int:id_ticket>/reasignar', methods=['POST'])
@jwt_required()
@login_required
def api_reasignar_ticket(id_ticket):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    nuevo_agente = data.get('id_agente')
    if not nuevo_agente:
        return jsonify({'ok': False, 'mensaje': 'id_agente requerido'}), 400
    descripcion = (data.get('descripcion') or '').strip()
    if not descripcion:
        return jsonify({'ok': False, 'mensaje': 'Descripción requerida'}), 400
    link_img = (data.get('link_img') or '').strip() or None
    # Verificar que el nuevo agente tiene nivel menor al usuario actual
    nivel_usuario = session['usuario'].get('nivel', 1)
    if nivel_usuario <= 1:
        return jsonify({'ok': False, 'mensaje': 'No tienes permiso para reasignar'}), 403
    try:
        reasignarTicket(id_ticket, int(nuevo_agente), descripcion, link_img)
        return jsonify({'ok': True})
    except Exception as e:
        print(f'Error al reasignar ticket: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error al reasignar ticket.'}), 500


@app.route('/api/ticket/<int:id_ticket>/calificar', methods=['POST'])
@jwt_required()
@login_required
def api_calificar_ticket(id_ticket):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    ticket = obtenerTicket(id_ticket)
    if not ticket:
        return jsonify({'ok': False, 'mensaje': 'Ticket no encontrado'}), 404
    if ticket['estado'] != 'resuelto':
        return jsonify({'ok': False, 'mensaje': 'Solo se puede calificar tickets resueltos'}), 400
    detalles = obtenerDetallesTicket(id_ticket)
    if any(d.get('activo') and d.get('calificacion_estrellas') for d in detalles):
        return jsonify({'ok': False, 'mensaje': 'El ticket ya fue calificado'}), 400
    estrellas  = data.get('estrellas')
    observacion = (data.get('observacion') or '').strip()
    if estrellas not in (1, 2, 3, 4, 5):
        return jsonify({'ok': False, 'mensaje': 'Calificación debe ser entre 1 y 5'}), 400
    guardarCalificacionTicket(ticket['id_detalle'], estrellas, observacion)
    return jsonify({'ok': True})


@app.route('/api/ticket/<int:id_ticket>/no-solucionado', methods=['POST'])
@jwt_required()
@login_required
def api_no_solucionado_ticket(id_ticket):
    """Califica y reabre el ticket (nuevo detalle con mismo agente, estado=en_progreso)."""
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    ticket = obtenerTicket(id_ticket)
    if not ticket:
        return jsonify({'ok': False, 'mensaje': 'Ticket no encontrado'}), 404
    if ticket['estado'] != 'resuelto':
        return jsonify({'ok': False, 'mensaje': 'Solo se puede reabrir tickets resueltos'}), 400
    estrellas  = data.get('estrellas')
    observacion = (data.get('observacion') or '').strip()
    if estrellas not in (1, 2, 3, 4, 5):
        return jsonify({'ok': False, 'mensaje': 'Calificación debe ser entre 1 y 5'}), 400
    descripcion = (data.get('descripcion') or '').strip()
    if not descripcion:
        return jsonify({'ok': False, 'mensaje': 'Descripción requerida para reabrir el ticket'}), 400
    link_img = (data.get('link_img') or '').strip() or None
    guardarCalificacionTicket(ticket['id_detalle'], estrellas, observacion)
    reabrirTicket(id_ticket, descripcion, link_img)
    return jsonify({'ok': True})


@app.route('/api/ticket/<int:id_ticket>/editar', methods=['POST'])
@jwt_required()
@login_required
def api_editar_ticket(id_ticket):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    ticket = obtenerTicket(id_ticket)
    if not ticket or ticket['estado'] != 'solicitado':
        return jsonify({'ok': False, 'mensaje': 'Solo se puede editar tickets en estado solicitado'}), 400
    titulo               = (data.get('titulo') or '').strip()
    tipo                 = data.get('tipo', '')
    id_aplicacion        = data.get('id_aplicacion', '')
    descripcion          = (data.get('descripcion') or '').strip()
    link_img_descripcion = (data.get('link_img_descripcion') or '').strip() or None
    if not titulo:
        return jsonify({'ok': False, 'mensaje': 'Título requerido'}), 400
    editarTicket(id_ticket, titulo, tipo, id_aplicacion, descripcion, link_img_descripcion)
    return jsonify({'ok': True})


@app.route('/api/ticket/<int:id_ticket>/cancelar', methods=['POST'])
@jwt_required()
@login_required
def api_cancelar_ticket(id_ticket):
    cancelarTicket(id_ticket)
    return jsonify({'ok': True})


@app.route('/api/ticket/<int:id_ticket>/asignar', methods=['POST'])
@jwt_required()
@login_required
def api_asignar_ticket(id_ticket):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    id_agente = data.get('id_agente')
    if not id_agente:
        return jsonify({'ok': False, 'mensaje': 'id_agente requerido'}), 400
    prioridad  = data.get('prioridad', 'media')
    intensidad = data.get('intensidad', 'media')
    sla_raw    = data.get('sla_horas')
    if sla_raw:
        sla_horas = int(sla_raw)
    else:
        id_app = data.get('id_aplicacion')
        app_data = obtenerAplicacion(int(id_app)) if id_app else None
        sla_horas = calcularSLA(prioridad, intensidad,
                                app_data['peso'] if app_data else 3,
                                app_data['participantes_promedio'] if app_data else 5)
    try:
        asignarTicket(id_ticket, int(id_agente), prioridad, intensidad, sla_horas)
        return jsonify({'ok': True})
    except ValueError:
        return jsonify({'ok': False, 'mensaje': 'Datos inválidos'}), 400
    except Exception as e:
        print(f'Error al asignar ticket: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error al asignar ticket.'}), 500


# ──────────────────────────────────────────────────────────────
#  API APLICACIONES CRUD
# ──────────────────────────────────────────────────────────────

@app.route('/api/aplicacion/nueva', methods=['POST'])
@jwt_required()
@login_required
def api_nueva_aplicacion():
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    nombre                = (data.get('nombre') or '').strip()
    peso                  = data.get('peso', 3)
    descripcion           = (data.get('descripcion') or '').strip()
    participantes_promedio = data.get('participantes_promedio', 5)
    if not nombre:
        return jsonify({'ok': False, 'mensaje': 'Nombre requerido'}), 400
    insertarAplicacion(nombre, peso, descripcion, participantes_promedio)
    return jsonify({'ok': True})


@app.route('/api/aplicacion/<int:id_aplicacion>/editar', methods=['POST'])
@jwt_required()
@login_required
def api_editar_aplicacion(id_aplicacion):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    app = obtenerAplicacion(id_aplicacion)
    if not app:
        return jsonify({'ok': False, 'mensaje': 'Aplicación no encontrada'}), 404
    nombre                = (data.get('nombre') or '').strip()
    peso                  = data.get('peso', 3)
    descripcion           = (data.get('descripcion') or '').strip()
    participantes_promedio = data.get('participantes_promedio', 5)
    if not nombre:
        return jsonify({'ok': False, 'mensaje': 'Nombre requerido'}), 400
    editarAplicacion(id_aplicacion, nombre, peso, descripcion, participantes_promedio)
    return jsonify({'ok': True})


@app.route('/api/aplicacion/<int:id_aplicacion>/eliminar', methods=['POST'])
@jwt_required()
@login_required
def api_eliminar_aplicacion(id_aplicacion):
    eliminarAplicacion(id_aplicacion)
    return jsonify({'ok': True})


@app.route('/api/aplicacion/<int:id_aplicacion>/toggle-estado', methods=['POST'])
@jwt_required()
@login_required
def api_toggle_estado_aplicacion(id_aplicacion):
    toggleEstadoAplicacion(id_aplicacion)
    return jsonify({'ok': True})


@app.route('/api/aplicaciones')
@jwt_required()
@login_required
def api_listar_aplicaciones():
    return jsonify({'ok': True, 'aplicaciones': listarAplicaciones()})


# ──────────────────────────────────────────────────────────────
#  API PROYECTOS CRUD
# ──────────────────────────────────────────────────────────────

@app.route('/api/proyecto/guardar', methods=['POST'])
@jwt_required()
@login_required
def api_guardar_proyecto():
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    try:
        from datetime import date
        ids_responsables = data.get('responsables', [])
        if not ids_responsables:
            return jsonify({'ok': False, 'mensaje': 'Al menos un responsable requerido'}), 400
        obj = Proyecto(
            nombre           = data['nombre'],
            ids_responsables = ids_responsables,
            estado           = 'en_revision',
            fecha_inicio     = date.today().strftime('%Y-%m-%d'),
            fecha_fin_plan   = data['fecha_fin_plan'],
            problematica     = (data.get('problematica') or '').strip() or None,
            justificacion    = (data.get('justificacion') or '').strip() or None,
            beneficios       = (data.get('beneficios') or '').strip() or None,
            descripcion      = data['descripcion'],
        )
        insertarProyecto(obj)
        return jsonify({'ok': True})
    except KeyError as e:
        return jsonify({'ok': False, 'mensaje': f'Campo requerido: {e}'}), 400
    except Exception as e:
        print(f'Error al guardar proyecto: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error al guardar proyecto.'}), 500


@app.route('/api/proyecto/<int:id_proyecto>/editar', methods=['POST'])
@jwt_required()
@login_required
def api_editar_proyecto(id_proyecto):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    try:
        from proyectoAD import actualizarProyectoCompleto
        ids_responsables = data.get('responsables', [])
        if not ids_responsables:
            return jsonify({'ok': False, 'mensaje': 'Al menos un responsable requerido'}), 400
        nombre         = data['nombre']
        estado         = data['estado']
        fecha_fin_plan = data['fecha_fin_plan']
        descripcion    = data['descripcion']
        actualizarProyectoCompleto(id_proyecto, nombre, ids_responsables, estado,
                                   fecha_fin_plan, descripcion)
        return jsonify({'ok': True})
    except KeyError as e:
        return jsonify({'ok': False, 'mensaje': f'Campo requerido: {e}'}), 400
    except Exception as e:
        print(f'Error al editar proyecto: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error al editar proyecto.'}), 500


@app.route('/api/proyecto/<int:id_proyecto>/avances/nuevo', methods=['POST'])
@jwt_required()
@login_required
def api_nuevo_avance(id_proyecto):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    proyecto = obtenerProyecto(id_proyecto)
    if not proyecto:
        return jsonify({'ok': False, 'mensaje': 'Proyecto no encontrado'}), 404
    from datetime import date
    id_autor = session['usuario']['id_usuario']
    resumen     = resumenActividadesPorProyecto()
    fila_actual = next((r for r in resumen
                        if (r['id_proyecto'] if isinstance(r, dict) else r[0]) == id_proyecto), None)
    if fila_actual:
        if isinstance(fila_actual, dict):
            valor_raw = fila_actual.get('porcentaje_avance_real')
        else:
            valor_raw = fila_actual[3]
        pct_calculado = round(float(valor_raw)) if valor_raw is not None else 0
    else:
        pct_calculado = 0
    if proyecto.get('estado') == 'completado':
        pct_calculado = 100
    hoy_db          = date.today().strftime('%Y-%m-%d')
    estado_salud    = data.get('estado_salud', 'ok')
    logros_periodo  = (data.get('logros_periodo') or '').strip()
    pendientes_next = (data.get('pendientes_next') or '').strip()
    insertarAvance(id_proyecto, id_autor, hoy_db, pct_calculado,
                   estado_salud, logros_periodo, pendientes_next)
    return jsonify({'ok': True, 'porcentaje': pct_calculado})


@app.route('/api/proyecto/<int:id_proyecto>/avances/eliminar/<int:id_avance>', methods=['POST'])
@jwt_required()
@login_required
def api_eliminar_avance_ruta(id_proyecto, id_avance):
    try:
        eliminarAvance(id_avance)
        return jsonify({'ok': True})
    except Exception as e:
        print(f"Error al eliminar avance: {e}")
        return jsonify({'ok': False, 'mensaje': 'No se pudo eliminar el avance.'})


# ──────────────────────────────────────────────────────────────
#  API ACTIVIDADES CRUD
# ──────────────────────────────────────────────────────────────

@app.route('/api/actividad/guardar', methods=['POST'])
@jwt_required()
@login_required
def api_guardar_actividad():
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    try:
        obj = Actividad(
            id_proyecto  = data['id_proyecto'],
            id_sprint    = data.get('id_sprint') or None,
            id_asignado  = data.get('id_asignado') or None,
            titulo       = data['titulo'],
            prioridad    = data['prioridad'],
            estado       = data['estado'],
            story_points = data.get('story_points') or 0,
        )
        insertarActividad(obj)
        return jsonify({'ok': True})
    except KeyError as e:
        return jsonify({'ok': False, 'mensaje': f'Campo requerido: {e}'}), 400
    except Exception as e:
        print(f'Error al guardar actividad: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error al guardar actividad.'}), 500


@app.route('/api/actividad/<int:id_actividad>/editar', methods=['POST'])
@jwt_required()
@login_required
def api_editar_actividad(id_actividad):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    try:
        from actividadAD import actualizarActividad, obtenerActividad
        if not obtenerActividad(id_actividad):
            return jsonify({'ok': False, 'mensaje': 'Actividad no encontrada'}), 404
        actualizarActividad(
            id_actividad = id_actividad,
            id_proyecto  = data['id_proyecto'],
            id_sprint    = data.get('id_sprint') or None,
            id_asignado  = data.get('id_asignado') or None,
            titulo       = data['titulo'],
            prioridad    = data['prioridad'],
            estado       = data['estado'],
            story_points = data.get('story_points') or 0,
        )
        return jsonify({'ok': True})
    except KeyError as e:
        return jsonify({'ok': False, 'mensaje': f'Campo requerido: {e}'}), 400
    except Exception as e:
        print(f'Error al editar actividad: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error al editar actividad.'}), 500


# ──────────────────────────────────────────────────────────────
#  APROBACION DE PROYECTOS (GERENTE)
# ──────────────────────────────────────────────────────────────

@app.route('/admin/aprobacion-proyectos')
@login_required
def aprobacion_proyectos():
    try:
        proyectos = listarProyectosEnRevision()
    except Exception as e:
        print(f'Error al listar proyectos en revision: {e}')
        abort(500)
    return render_template('aprobacionProyectos.html', proyectos=proyectos)


@app.route('/api/proyecto/<int:id_proyecto>/aprobar', methods=['POST'])
@jwt_required()
@login_required
def api_aprobar_proyecto(id_proyecto):
    try:
        from datetime import date
        # Obtener datos del proyecto antes de aprobar para calcular sprints
        proyecto = obtenerProyecto(id_proyecto)
        aprobarProyecto(id_proyecto)
        # Generar sprints automáticamente si no los tiene ya
        if proyecto and proyecto.get('fecha_inicio') and proyecto.get('fecha_fin_plan'):
            from proyectoAD import listarSprintsPorProyecto
            sprints_existentes = listarSprintsPorProyecto(id_proyecto)
            if not sprints_existentes:
                generarSprintsProyecto(
                    id_proyecto,
                    proyecto['fecha_inicio'],
                    proyecto['fecha_fin_plan']
                )
        return jsonify({'ok': True})
    except Exception as e:
        print(f'Error al aprobar proyecto: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error al aprobar el proyecto.'})


@app.route('/api/proyecto/<int:id_proyecto>/sprints')
@jwt_required()
@login_required
def api_sprints_por_proyecto(id_proyecto):
    """Devuelve los sprints de un proyecto en JSON para el combo dinámico."""
    try:
        sprints = listarSprintsPorProyecto(id_proyecto)
        resultado = [
            {
                'id_sprint': s['id_sprint'],
                'nombre':    s['nombre'],
                'estado':    s['estado'],
                'fecha_inicio': str(s['fecha_inicio']),
                'fecha_fin':    str(s['fecha_fin']),
            }
            for s in sprints
        ]
        return jsonify({'ok': True, 'sprints': resultado})
    except Exception as e:
        print(f'Error al obtener sprints: {e}')
        return jsonify({'ok': False, 'sprints': []})


@app.route('/api/proyecto/<int:id_proyecto>/rechazar', methods=['POST'])
@jwt_required()
@login_required
def api_rechazar_proyecto(id_proyecto):
    try:
        rechazarProyecto(id_proyecto)
        return jsonify({'ok': True})
    except Exception as e:
        print(f'Error al rechazar proyecto: {e}')
        return jsonify({'ok': False, 'mensaje': 'Error al rechazar el proyecto.'})


@app.route('/api/usuario/nuevo', methods=['POST'])
@jwt_required()
@login_required
def api_nuevo_usuario():
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    nombre_completo = (data.get('nombre_completo') or '').strip()
    correo = (data.get('correo') or '').strip()
    password = data.get('password') or ''
    id_rol = data.get('id_rol')
    if not nombre_completo or not correo or not password or not id_rol:
        return jsonify({'ok': False, 'mensaje': 'nombre_completo, correo, password e id_rol son requeridos'}), 400
    ok, result = insertarUsuario(
        nombre_completo=nombre_completo,
        correo=correo,
        password=password,
        id_rol=id_rol,
        nivel=data.get('nivel', 1),
        apellido=(data.get('apellido') or '').strip() or None,
        edad=data.get('edad'),
        dni=(data.get('dni') or '').strip() or None,
        direccion=(data.get('direccion') or '').strip() or None,
        foto_url=(data.get('foto_url') or '').strip() or None,
    )
    if ok:
        return jsonify({'ok': True, 'id_usuario': result})
    return jsonify({'ok': False, 'mensaje': result}), 400


# ──────────────────────────────────────────────────────────────
#  CONFIGURACIÓN
# ──────────────────────────────────────────────────────────────

@app.route('/configuracion')
@login_required
def configuracion():
    roles = listarRoles()
    aplicaciones = listarAplicaciones()
    return render_template('configuracion.html', roles=roles, aplicaciones=aplicaciones)


# ──────────────────────────────────────────────────────────────
#  ROLES — Formularios (páginas)
# ──────────────────────────────────────────────────────────────

@app.route('/rol/nuevo')
@login_required
def form_nuevo_rol():
    return render_template('formularioRol.html')


@app.route('/rol/<int:id_rol>/editar')
@login_required
def form_editar_rol(id_rol):
    roles = listarRoles()
    rol = next((r for r in roles if r['id_rol'] == id_rol), None)
    if not rol:
        abort(404)
    return render_template('formularioRol.html', rol=rol)


# ──────────────────────────────────────────────────────────────
#  API ROLES CRUD
# ──────────────────────────────────────────────────────────────

@app.route('/api/roles')
@jwt_required()
@login_required
def api_listar_roles():
    return jsonify({'ok': True, 'roles': listarRoles()})


@app.route('/api/rol', methods=['POST'])
@jwt_required()
@login_required
def api_crear_rol():
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    nombre = (data.get('nombre') or '').strip()
    descripcion = (data.get('descripcion') or '').strip()
    if not nombre:
        return jsonify({'ok': False, 'mensaje': 'Nombre requerido'}), 400
    id_rol = insertarRol(nombre, descripcion)
    return jsonify({'ok': True, 'id_rol': id_rol})


@app.route('/api/rol/<int:id_rol>', methods=['PUT'])
@jwt_required()
@login_required
def api_actualizar_rol(id_rol):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    nombre = (data.get('nombre') or '').strip()
    descripcion = (data.get('descripcion') or '').strip()
    if not nombre:
        return jsonify({'ok': False, 'mensaje': 'Nombre requerido'}), 400
    actualizarRol(id_rol, nombre, descripcion)
    return jsonify({'ok': True})


@app.route('/api/rol/<int:id_rol>', methods=['DELETE'])
@jwt_required()
@login_required
def api_eliminar_rol(id_rol):
    ok = eliminarRol(id_rol)
    if ok:
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'mensaje': 'No se puede eliminar: el rol tiene usuarios asignados.'})


# ──────────────────────────────────────────────────────────────
#  PERMISO FORM ROUTES
# ──────────────────────────────────────────────────────────────

@app.route('/permiso/nuevo')
@login_required
def form_nuevo_permiso():
    from rolAD import listarRoles
    return render_template('formularioPermiso.html', roles=listarRoles())


@app.route('/permiso/<int:id_rol_permiso>/editar')
@login_required
def form_editar_permiso(id_rol_permiso):
    permiso = obtenerPermiso(id_rol_permiso)
    if not permiso:
        abort(404)
    return render_template('formularioPermiso.html', permiso=permiso)


# ──────────────────────────────────────────────────────────────
#  API PERMISOS CRUD
# ──────────────────────────────────────────────────────────────

@app.route('/api/rol/<int:id_rol>/permisos')
@jwt_required()
@login_required
def api_permisos_por_rol(id_rol):
    return jsonify({'ok': True, 'permisos': listarPermisos(id_rol)})


@app.route('/api/permisos')
@jwt_required()
@login_required
def api_listar_permisos():
    id_rol = request.args.get('id_rol', type=int)
    return jsonify({'ok': True, 'permisos': listarPermisos(id_rol)})


@app.route('/api/permiso', methods=['POST'])
@jwt_required()
@login_required
def api_crear_permiso():
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    nombre = (data.get('nombre') or '').strip()
    nivel = data.get('nivel', 1)
    id_rol = data.get('id_rol')
    if not nombre or not id_rol:
        return jsonify({'ok': False, 'mensaje': 'Nombre e id_rol requeridos'}), 400
    id_rol_permiso = insertarPermiso(nombre, nivel, id_rol)
    return jsonify({'ok': True, 'id_rol_permiso': id_rol_permiso})


@app.route('/api/permiso/<int:id_rol_permiso>', methods=['PUT'])
@jwt_required()
@login_required
def api_actualizar_permiso(id_rol_permiso):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'mensaje': 'JSON requerido'}), 400
    nombre = (data.get('nombre') or '').strip()
    nivel = data.get('nivel', 1)
    id_rol = data.get('id_rol')
    if not nombre or not id_rol:
        return jsonify({'ok': False, 'mensaje': 'Nombre e id_rol requeridos'}), 400
    actualizarPermiso(id_rol_permiso, nombre, nivel, id_rol)
    return jsonify({'ok': True})


@app.route('/api/permiso/<int:id_rol_permiso>', methods=['DELETE'])
@jwt_required()
@login_required
def api_eliminar_permiso(id_rol_permiso):
    eliminarPermiso(id_rol_permiso)
    return jsonify({'ok': True})

@app.route('/api/chatbot/contexto')
@jwt_required()
@login_required
def api_chatbot_contexto():
    """Devuelve datos reales del usuario según su rol para alimentar el chatbot."""
    id_usuario = session['usuario']['id_usuario']
    nivel      = session['usuario'].get('nivel', 1)
    rol_nombre = session['usuario'].get('rol_nombre', '').lower()

    todos_tickets = listarTickets()

    # ── SOPORTE (nivel 4) → solo ve sus tickets asignados para resolver
    if 'soporte' in rol_nombre or nivel == 4:
        asignados = [t for t in todos_tickets
                     if t.get('id_agente') == id_usuario and t['estado'] == 'en_progreso']
        resueltos = [t for t in todos_tickets
                     if t.get('id_agente') == id_usuario and t['estado'] == 'resuelto']
        return jsonify({
            'ok': True,
            'resumen': {
                'tickets_asignados'     : len(asignados),
                'tickets_resueltos'     : len(resueltos),
                'proyectos_activos'     : 0,
                'actividades_pendientes': 0,
                'tickets_abiertos'      : 0,
            },
            'tickets_asignados': [
                {'id': t['id_ticket'], 'titulo': t['titulo'], 'prioridad': t.get('prioridad', 'media')}
                for t in asignados[:5]
            ],
            'tickets_abiertos'      : [],
            'proyectos'             : [],
            'actividades_pendientes': [],
        })

    # ── PROGRAMADOR (nivel 3) → solo ve sus actividades y proyectos asignados
    if 'programador' in rol_nombre or nivel == 3:
        todas_acts    = listarActividades()
        mis_acts_pend = [a for a in todas_acts
                         if a.get('id_asignado') == id_usuario
                         and a['estado'] in ('por_hacer', 'en_progreso', 'bloqueado')]
        todos_proyectos = listarProyectos()
        try:
            from conexion import obtenerconexion as _conn
            c = _conn()
            with c:
                with c.cursor() as cur:
                    cur.execute("SELECT id_proyecto FROM asignado WHERE id_usuario = %s", (id_usuario,))
                    ids_proy = {r['id_proyecto'] for r in cur.fetchall()}
        except Exception:
            ids_proy = set()
        mis_proyectos = [p for p in todos_proyectos if p['id_proyecto'] in ids_proy]
        return jsonify({
            'ok': True,
            'resumen': {
                'tickets_asignados'     : 0,
                'tickets_resueltos'     : 0,
                'tickets_abiertos'      : 0,
                'proyectos_activos'     : len(mis_proyectos),
                'actividades_pendientes': len(mis_acts_pend),
            },
            'tickets_abiertos'      : [],
            'tickets_asignados'     : [],
            'proyectos': [
                {'nombre': p['nombre'], 'estado': p.get('estado_bd', ''),
                 'total_acts': p.get('total_acts', 0), 'completadas': p.get('completadas', 0)}
                for p in mis_proyectos[:5]
            ],
            'actividades_pendientes': [
                {'titulo': a['titulo'], 'estado': a['estado'], 'prioridad': a.get('prioridad', 'media')}
                for a in mis_acts_pend[:5]
            ],
        })

    # ── ADMIN (nivel 5) → ve todo
    mis_tickets   = [t for t in todos_tickets
                     if t.get('id_solicitante') == id_usuario or t.get('id_agente') == id_usuario]
    abiertos  = [t for t in mis_tickets if t['estado'] in ('solicitado', 'en_progreso')]
    resueltos = [t for t in mis_tickets if t['estado'] == 'resuelto']
    asignados = [t for t in mis_tickets if t.get('id_agente') == id_usuario and t['estado'] == 'en_progreso']

    todos_proyectos = listarProyectos()
    try:
        from conexion import obtenerconexion as _conn
        c = _conn()
        with c:
            with c.cursor() as cur:
                cur.execute("""
                    SELECT id_proyecto FROM asignado WHERE id_usuario = %s
                    UNION
                    SELECT id_proyecto FROM proyectos WHERE id_Stakeholder = %s
                """, (id_usuario, id_usuario))
                ids_proy = {r['id_proyecto'] for r in cur.fetchall()}
    except Exception:
        ids_proy = set()
    mis_proyectos = [p for p in todos_proyectos if p['id_proyecto'] in ids_proy]

    todas_acts    = listarActividades()
    mis_acts_pend = [a for a in todas_acts
                     if a.get('id_asignado') == id_usuario
                     and a['estado'] in ('por_hacer', 'en_progreso', 'bloqueado')]

    return jsonify({
        'ok': True,
        'resumen': {
            'tickets_abiertos'      : len(abiertos),
            'tickets_resueltos'     : len(resueltos),
            'tickets_asignados'     : len(asignados),
            'proyectos_activos'     : len(mis_proyectos),
            'actividades_pendientes': len(mis_acts_pend),
        },
        'tickets_abiertos': [
            {'id': t['id_ticket'], 'titulo': t['titulo'],
             'estado': t['estado'], 'prioridad': t.get('prioridad', 'media')}
            for t in abiertos[:5]
        ],
        'tickets_asignados': [
            {'id': t['id_ticket'], 'titulo': t['titulo'], 'prioridad': t.get('prioridad', 'media')}
            for t in asignados[:5]
        ],
        'proyectos': [
            {'nombre': p['nombre'], 'estado': p.get('estado_bd', ''),
             'total_acts': p.get('total_acts', 0), 'completadas': p.get('completadas', 0)}
            for p in mis_proyectos[:5]
        ],
        'actividades_pendientes': [
            {'titulo': a['titulo'], 'estado': a['estado'], 'prioridad': a.get('prioridad', 'media')}
            for a in mis_acts_pend[:5]
        ],
    })


# ──────────────────────────────────────────────────────────────
#  ARRANQUE
# ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    app.run(debug=True)